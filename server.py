import paho.mqtt.client as mqtt
import openpyxl as xl
from datetime import timedelta, datetime, tzinfo, timezone
import time
import os
def on_connect(client, userdata, flags, rc):
    print("Connected with result code "+str(rc))
    
    # string = " " + b + " "+c +" "+ d+ " "+e +" "+ f
    # client.publish("wxstation/wind_speed",string)

filename1 = os.getcwd() + "\\18-07-20.xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active

client = mqtt.Client()
client.on_connect = on_connect
client.connect("broker.hivemq.com", 1883, 60)
client.loop_start()

for i in range (2, 120):  
    # reading cell value from source excel file 
        d = str(ws2.cell(row = i, column = 4).value)
        #print(c)
        a = str(ws2.cell(row = i, column = 1).value)
        b = str(ws2.cell(row = i, column = 2).value)
        c = str(ws2.cell(row = i, column = 3).value)
        e = str(ws2.cell(row = i, column = 5).value)
        f = str(ws2.cell(row = i, column = 6).value)
        client.publish("pibpump/a",a + " " + d + " "  + b + " " + c + " " + e + " " + f)
        client.publish("pibpump/b",a + " " + d + " "  + b + " " + c + " " + e + " " + f)
        client.publish("pibpump/c",a + " " + d + " "  + b + " " + c + " " + e + " " + f)
        

        time.sleep(1)